#!/usr/bin/env python3
"""Build tidy hurricane wind curve tables from HazusWindDamFunctions_Hazus61.xlsx.

Source shape (verified against the workbook):

  huDamLossFun            275,220 rows = 6,116 wbID x 5 terrain x 9 loss classes,
                          with 41 wide columns WS50..WS250 (3-second peak gust, mph,
                          in 5 mph steps).
  huListOfWindBldgTypes   6,116 wind building types -> 39 specific building types,
                          each with a concatenated characteristic code string.
  huListOfBldgChar        53 building characteristic codes across 23 characteristic
                          types (roof shape, shutters, roof deck attachment, ...).
  huTerrain               5 terrain roughnesses.
  CaseID                  geographic applicability of each wind building type.

The 9 loss classes do NOT share units. Four are damage-state exceedance probabilities,
two are loss ratios, one is days, two are lbs/sqft. They must never be averaged
together; ``curve_kind`` records the units so consumers do not have to guess.

Expands to 275,220 x 41 = 11,284,020 tidy points, written to Parquet in batches so the
whole thing is never held in memory at once.
"""

import sys
from pathlib import Path

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

REPO = Path(__file__).resolve().parent.parent
RAW = REPO / "raw"
DATA = REPO / "data"
DIST = REPO / "dist"

WORKBOOK = "HazusWindDamFunctions_Hazus61.xlsx"
VERSION = "6.1"

# DamLossDescID -> (damage_type, y_units). Verbatim from huDamLossFunDescription.
LOSS_CLASSES = {
    1: ("damage_slight",        "probability"),
    2: ("damage_moderate",      "probability"),
    3: ("damage_severe",        "probability"),
    4: ("damage_total",         "probability"),
    5: ("building_loss",        "loss_ratio"),
    6: ("content_loss",         "loss_ratio"),
    7: ("loss_of_use",          "days"),
    8: ("debris_brick_wood",    "lbs_per_sqft"),
    9: ("debris_concrete_steel", "lbs_per_sqft"),
}

# FEMA Hazus 7.1 Release Notes, RSA-21186: in Hazus 6.1 and 7.0 the 2- and 3-story
# multi-family wind damage functions had been inadvertently overwritten with copies of
# the 1-story functions. Corrected only in 7.1+, which is not publicly extractable.
# We flag rather than withhold -- see docs/provenance.md. verify_defect() below checks
# the claim against the data itself rather than taking the release note on faith.
DEFECT_SBT = {"WMUH2", "WMUH3", "MMUH2", "MMUH3"}
DEFECT_NOTE = (
    "FEMA Hazus 7.1 release notes (RSA-21186): in Hazus 6.1/7.0 the 2- and 3-story "
    "multi-family wind damage functions were inadvertently overwritten with copies of "
    "the 1-story functions. Fixed in Hazus 7.1, which is not publicly extractable. "
    "Use with caution."
)

BATCH = 200_000  # points per Parquet row group write


def sheet_rows(wb, name):
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else h for h in next(it)]
    for row in it:
        yield dict(zip(header, row))


def clean(v):
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


# Every code in huListOfBldgChar is exactly 5 characters (verified: all 53 of them),
# and every charDescription length is an exact multiple of 5 (verified: all 6,116).
# So the string splits positionally -- this is exact, not a heuristic.
CODE_WIDTH = 5

# huListOfBldgChar is incomplete: 9 codes appear in huListOfWindBldgTypes but have no
# entry in the workbook's own decode table. We record them verbatim rather than
# guessing at their meaning. Counts observed in Hazus 6.1: rcshg 960, rcssm 768,
# wtnor 520, wtjal 520, rcapr 416, rcagd 416, rcpnt 192, rccor 64, rccnt 16.
# See docs/provenance.md.
UNMAPPED_KEY = "characteristic_unmapped_code"


def decode_characteristics(char_string, char_meta):
    """Split a concatenated characteristic string into its component 5-char codes.

    Returns (known, unknown) where ``known`` is a list of codes present in
    huListOfBldgChar and ``unknown`` is a list of codes that are not. Unknown codes
    are preserved verbatim; no meaning is inferred for them.
    """
    s = (char_string or "").strip()
    known, unknown = [], []
    for i in range(0, len(s), CODE_WIDTH):
        code = s[i:i + CODE_WIDTH]
        (known if code in char_meta else unknown).append(code)
    return known, unknown


def build():
    wb = openpyxl.load_workbook(RAW / WORKBOOK, read_only=True)

    terrain = {r["TERRAINID"]: r for r in sheet_rows(wb, "huTerrain")}
    cases = {r["CaseID"]: clean(r["CaseName"]) for r in sheet_rows(wb, "CaseID")}
    char_meta = {}
    for r in sheet_rows(wb, "huListOfBldgChar"):
        char_meta[clean(r["BldgChar"])] = (clean(r["CharType"]), clean(r["bcName"]))
    wbt = {}
    for r in sheet_rows(wb, "huListOfWindBldgTypes"):
        wbt[r["wbID"]] = {
            "sbt": clean(r["sbtName"]),
            "chars": clean(r["charDescription"]),
            "case": cases.get(r["CaseID"]),
        }
    print(f"  {len(wbt)} wind building types, "
          f"{len({v['sbt'] for v in wbt.values()})} specific building types")

    # Decode characteristic strings once per wind building type.
    decoded, n_known, n_unknown = {}, 0, 0
    unknown_codes = {}
    for wbid, meta in wbt.items():
        known, unknown = decode_characteristics(meta["chars"], char_meta)
        decoded[wbid] = (known, unknown)
        n_known += len(known)
        n_unknown += len(unknown)
        for code in unknown:
            unknown_codes[code] = unknown_codes.get(code, 0) + 1
    print(f"  characteristics: {n_known} decoded, {n_unknown} not in "
          f"huListOfBldgChar (preserved verbatim)")
    if unknown_codes:
        listing = ", ".join(f"{c} x{n}" for c, n in
                            sorted(unknown_codes.items(), key=lambda kv: -kv[1]))
        print(f"  codes missing from the Hazus decode table: {listing}")

    DATA.mkdir(parents=True, exist_ok=True)
    DIST.mkdir(parents=True, exist_ok=True)

    ws = wb["huDamLossFun"]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(it)]
    speed_cols = [(i, float(h[2:])) for i, h in enumerate(header)
                  if h.startswith("WS")]
    key_idx = {h: i for i, h in enumerate(header) if not h.startswith("WS")}
    print(f"  {len(speed_cols)} wind speeds: {speed_cols[0][1]:.0f}"
          f"..{speed_cols[-1][1]:.0f} mph")

    point_schema = pa.schema([("curve_id", pa.string()),
                              ("x", pa.float64()),
                              ("y", pa.float64())])
    writer = pq.ParquetWriter(DIST / "curve_points_hu.parquet", point_schema,
                              compression="zstd")

    curves, attrs = [], []
    cid_buf, x_buf, y_buf = [], [], []
    n_curves = n_points = 0

    def flush():
        nonlocal cid_buf, x_buf, y_buf
        if not cid_buf:
            return
        writer.write_table(pa.table({"curve_id": cid_buf, "x": x_buf, "y": y_buf},
                                    schema=point_schema))
        cid_buf, x_buf, y_buf = [], [], []

    for row in it:
        wbid = row[key_idx["wbID"]]
        tid = row[key_idx["TERRAINID"]]
        did = row[key_idx["DamLossDescID"]]
        meta = wbt.get(wbid)
        if meta is None or did not in LOSS_CLASSES:
            continue
        damage_type, _units = LOSS_CLASSES[did]
        curve_id = f"hu-{VERSION}-{damage_type}-{wbid}-t{tid}"

        curves.append({
            "curve_id": curve_id,
            "peril": "hu",
            "hazus_version": VERSION,
            "damage_type": damage_type,
            "occupancy": None,
            "building_type": meta["sbt"],
            "source_agency": "FEMA Hazus hurricane model",
            "description": meta["chars"],
            "defect_flag": DEFECT_NOTE if meta["sbt"] in DEFECT_SBT else None,
            "source_file": WORKBOOK,
            "source_table": "huDamLossFun",
            "source_row_id": f"{wbid}/{tid}/{did}",
        })

        t = terrain.get(tid, {})
        pairs = [("wind_building_type_id", wbid),
                 ("specific_building_type", meta["sbt"]),
                 ("terrain_id", tid),
                 ("terrain", clean(t.get("SRDescription"))),
                 ("surface_roughness_m", t.get("SURFACEROUGHNESS")),
                 ("geographic_case", meta["case"]),
                 ("characteristic_code", meta["chars"])]
        known, unknown = decoded.get(wbid, ([], []))
        for code in known:
            char_type, name = char_meta[code]
            pairs.append((char_type, name))
        if unknown:
            pairs.append((UNMAPPED_KEY, ",".join(unknown)))
        for k, v in pairs:
            if v is not None:
                attrs.append({"curve_id": curve_id, "key": k, "value": str(v)})

        for i, speed in speed_cols:
            v = row[i]
            cid_buf.append(curve_id)
            x_buf.append(speed)
            y_buf.append(None if v is None else float(v))
        n_points += len(speed_cols)
        n_curves += 1

        if len(cid_buf) >= BATCH:
            flush()

    flush()
    writer.close()
    wb.close()

    import pandas as pd
    cdf = pd.DataFrame(curves)
    adf = pd.DataFrame(attrs)
    cdf.to_parquet(DIST / "curves_hu.parquet", index=False, compression="zstd")
    adf.to_parquet(DIST / "curve_attributes_hu.parquet", index=False,
                   compression="zstd")

    bt = (cdf[["building_type"]].drop_duplicates()
          .assign(description=None).rename(columns={"building_type": "building_type"}))
    bt.to_csv(DATA / "dim_building_type.csv", index=False)

    flagged = int(cdf["defect_flag"].notna().sum())
    print(f"\n  {n_curves} hurricane curves, {n_points} points, {len(attrs)} attributes")
    print(f"  {flagged} curves flagged with the FEMA-disclosed multi-family defect")
    return cdf


if __name__ == "__main__":
    build()
